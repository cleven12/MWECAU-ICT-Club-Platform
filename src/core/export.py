"""
PDF generation and export utilities
"""
from io import BytesIO
import logging

logger = logging.getLogger(__name__)


class PDFGenerator:
    """PDF generation utilities"""
    
    @staticmethod
    def generate_user_report(user) -> bytes:
        """Generate PDF report for user"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#2c3e50'),
                spaceAfter=30
            )
            elements.append(Paragraph('User Report', title_style))
            elements.append(Spacer(1, 12))
            
            # User info
            user_data = [
                ['Field', 'Value'],
                ['Name', user.full_name],
                ['Registration Number', user.reg_number],
                ['Email', user.email],
                ['Department', str(user.department)],
                ['Course', str(user.course)],
                ['Status', 'Approved' if user.is_approved else 'Pending'],
            ]
            
            table = Table(user_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            return buffer.getvalue()
        except ImportError:
            logger.error("ReportLab not installed for PDF generation")
            return None
    
    @staticmethod
    def generate_event_list_pdf(events) -> bytes:
        """Generate PDF list of events"""
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            elements.append(Paragraph('Event List', styles['Heading1']))
            elements.append(Spacer(1, 12))
            
            # Events table
            event_data = [['Title', 'Date', 'Location', 'Department']]
            for event in events:
                event_data.append([
                    event.title,
                    event.event_date.strftime('%Y-%m-%d'),
                    event.location,
                    str(event.department)
                ])
            
            table = Table(event_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            elements.append(table)
            doc.build(elements)
            
            return buffer.getvalue()
        except ImportError:
            logger.error("ReportLab not installed for PDF generation")
            return None


class ExportHelper:
    """Data export utilities"""
    
    @staticmethod
    def export_to_csv(data: list, filename: str = 'export.csv'):
        """Export data to CSV"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.DictWriter(response, fieldnames=data[0].keys() if data else [])
        writer.writeheader()
        writer.writerows(data)
        
        return response
    
    @staticmethod
    def export_to_json(data: list, filename: str = 'export.json'):
        """Export data to JSON"""
        import json
        from django.http import HttpResponse
        
        response = HttpResponse(
            json.dumps(data, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @staticmethod
    def export_to_excel(data: list, filename: str = 'export.xlsx'):
        """Export data to Excel"""
        try:
            import openpyxl
            from django.http import HttpResponse
            
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Write headers
            if data:
                for col, key in enumerate(data[0].keys(), 1):
                    ws.cell(row=1, column=col, value=key)
                
                # Write data
                for row, item in enumerate(data, 2):
                    for col, value in enumerate(item.values(), 1):
                        ws.cell(row=row, column=col, value=value)
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
        except ImportError:
            logger.error("openpyxl not installed for Excel export")
            return None


class DataImport:
    """Data import utilities"""
    
    @staticmethod
    def import_users_from_csv(csv_file):
        """Import users from CSV file"""
        import csv
        from accounts.models import CustomUser, Department, Course
        
        imported = 0
        errors = []
        
        try:
            for row in csv.DictReader(csv_file):
                try:
                    department = Department.objects.get(name=row['department'])
                    course = Course.objects.get(code=row['course'])
                    
                    CustomUser.objects.create_user(
                        reg_number=row['reg_number'],
                        email=row['email'],
                        password=row.get('password', 'TempPassword123!'),
                        full_name=row['full_name'],
                        department=department,
                        course=course,
                    )
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {imported + 1}: {str(e)}")
        except Exception as e:
            logger.error(f"Failed to import users: {str(e)}")
        
        return {
            'imported': imported,
            'errors': errors
        }
    
    @staticmethod
    def import_events_from_csv(csv_file):
        """Import events from CSV file"""
        import csv
        from core.models import Event
        from accounts.models import Department
        from datetime import datetime
        
        imported = 0
        errors = []
        
        try:
            for row in csv.DictReader(csv_file):
                try:
                    department = Department.objects.get(name=row['department'])
                    
                    Event.objects.create(
                        title=row['title'],
                        description=row.get('description', ''),
                        event_date=datetime.fromisoformat(row['event_date']),
                        location=row['location'],
                        department=department,
                    )
                    imported += 1
                except Exception as e:
                    errors.append(f"Row {imported + 1}: {str(e)}")
        except Exception as e:
            logger.error(f"Failed to import events: {str(e)}")
        
        return {
            'imported': imported,
            'errors': errors
        }
